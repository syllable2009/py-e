from openpyxl import Workbook, load_workbook

# === 1. 新建 Excel 并写入数据 ===
def create_excel(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    # 写入表头
    ws.append(["ID", "Name", "Department", "Salary"])

    # 写入数据
    data = [
        [1, "Alice", "HR", 5000],
        [2, "Bob", "IT", 7000],
        [3, "Charlie", "Finance", 6000],
        [4, "Diana", "IT", 7500],
    ]
    for row in data:
        ws.append(row)

    wb.save(filename)
    print(f"✅ Excel 文件 '{filename}' 已创建并保存。")

# === 2. 读取 Excel 并搜索内容 ===
def search_in_excel(filename, keyword, column_name="Name"):
    wb = load_workbook(filename)
    ws = wb.active

    # 获取表头（第一行）
    headers = [cell.value for cell in ws[1]]
    try:
        col_index = headers.index(column_name) + 1  # openpyxl 列索引从 1 开始
    except ValueError:
        print(f"❌ 列 '{column_name}' 不存在！")
        return []

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if row[col_index - 1] and keyword.lower() in str(row[col_index - 1]).lower():
            results.append(row)

    return results

# === 使用示例 ===
if __name__ == "__main__":
    filename = "/Users/jiaxiaopeng/ppt/employees.xlsx"

    # 创建文件
    create_excel(filename)

    # 搜索名字包含 "a" 的员工（不区分大小写）
    matches = search_in_excel(filename, keyword="a", column_name="Name")
    print("\n🔍 搜索结果（Name 列包含 'a'）:")
    for match in matches:
        print(match)

    # 搜索 IT 部门
    it_staff = search_in_excel(filename, keyword="IT", column_name="Department")
    print("\n🔍 IT 部门员工:")
    for staff in it_staff:
        print(staff)