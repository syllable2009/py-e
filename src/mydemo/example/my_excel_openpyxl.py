import os
from openpyxl import Workbook, load_workbook
from typing import List, Any, Optional, Tuple

from mydemo.exception.business_exception import BusinessException
from mydemo.exception.business_exception_constant import ExceptionCode


# 在 Python 3 中操作 Excel 文件，最常用、功能最强大的库主要有两个：
# openpyxl：用于读写 .xlsx（Excel 2007 及以后）格式文件，不支持 .xls。
# pandas + openpyxl / xlsxwriter：适合数据分析、批量处理。

class ExcelOpenpyxl:
    def __init__(self, file_name: str):
        """
        初始化 Excel 文件管理器
        :param filename: Excel 文件路径（支持 .xlsx）
        """
        self.file_name = file_name
        if os.path.exists(file_name):
            self._wb: Optional[Workbook] = load_workbook(self.file_name)
        else:
            # 文件不存在：创建新工作簿
            self._wb = Workbook()
            ws = self._wb.active
            ws.title = 'Sheet1'
            self._wb.save(file_name)
            print(f"🆕 创建新 Excel 文件 '{file_name}'，并添加 Sheet: '{ws.title}'")

    def _load_excel(self):
        """确保工作簿已加载或创建"""
        if self._wb is not None:
            return
        if os.path.exists(self.filename):
            self._wb = load_workbook(self.filename)
        else:
            raise BusinessException(ExceptionCode.NOT_FOUND.name, ExceptionCode.NOT_FOUND.code)

    def add_sheet(self, sheet_name: str):
        if sheet_name in self._wb.sheetnames:
            print(f"⚠️ Sheet '{sheet_name}' 已存在于文件 '{self.file_name}' 中，跳过创建。")
            return
        else:
            # 新增 sheet
            ws = self._wb.create_sheet(title=sheet_name)
            ws.title = sheet_name
            self._wb.save(self.file_name)
            print(f"➕ 在现有文件 '{self.file_name}' 中新增 Sheet: '{sheet_name}'")

    def get_sheet(self, sheet_name: str):
        """确保工作簿已加载或创建"""
        if sheet_name in self._wb.sheetnames:
            return self._wb[sheet_name]
        else:
            raise BusinessException(ExceptionCode.NOT_FOUND.name, ExceptionCode.NOT_FOUND.code)

    def set_header(self, sheet_name: str, header: List[str]):
        ws = self.get_sheet(sheet_name)
        if not ws:
            print(f"❌ sheet '{sheet_name}' 不存在！跳过设置sheet表头。")
        # openpyxl 的行和列索引从 1 开始（不是 0）
        if ws.max_row == 1 and ws.cell(1, 1).value is None:
            row_index = 1
            for col_index, value in enumerate(header, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
            self._wb.save(self.file_name)
        else:
            ws.append(header)
            self._wb.save(self.file_name)
            print(f"⚠️ Sheet '{sheet_name}' 已存在！跳过设置sheet表头。")

    # 如果存在row为赋值，否则为设置row的值
    def set_row(self, sheet_name: str, row_index: Optional[int], data: List[str]):
        ws = self.get_sheet(sheet_name)
        if not ws:
            print(f"❌ sheet '{sheet_name}' 不存在！")
            return
        if row_index is None:
            ws.append(data)
            self._wb.save(self.file_name)
            print(f"⚠️ Sheet '{sheet_name}' append data。")
        else:
            for col_index, value in enumerate(data, start=1):
                ws.cell(row=row_index, column=col_index, value=value)
            self._wb.save(self.file_name)
            print(f"⚠️ Sheet '{sheet_name}' 设置 row '{row_index}' data。")

    def search_in_column(self, sheet_name: str, column_name, keyword) -> List[Tuple[int, Tuple[Any, ...]]]:
        ws = self.get_sheet(sheet_name)
        if not ws:
            print(f"❌ sheet '{sheet_name}' 不存在！")
            return None
        # 获取表头（第一行）
        # 读取表头（第1行）
        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            print("❌ 表头为空或无效！")
            return None
        # print(headers)
        try:
            header_lower = [str(h).lower() if h is not None else '' for h in headers]
            target_col_lower = str(column_name).lower()
            col_offset = header_lower.index(target_col_lower)
            col_index = col_offset + 1  # openpyxl 列索引从 1 开始
        except ValueError:
            print(f"❌ 列 '{column_name}' 不存在！")
            return None
        # print(col_index)
        results = []
        # 从第2行开始遍历（跳过表头），同时获取行号
        for row_idx in range(2, ws.max_row + 1):
            row = tuple(ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1))
            cell_value = row[col_offset]
            # 安全地进行字符串包含匹配
            if cell_value is not None:
                try:
                    if keyword.lower() in str(cell_value).lower():
                        results.append((row_idx, row))
                except Exception:
                    # 某些特殊对象（如 datetime）转 str 可能异常，跳过
                    continue
        return results


if __name__ == "__main__":
    file_path = "/Users/jiaxiaopeng/ppt/测试.xlsx"
    # 不存在会默认创建一个
    eo = ExcelOpenpyxl(file_path)
    eo._load_excel()
    # data = ["name", "age"]
    # eo.add_sheet('我的文档')
    # eo.set_row("我的文档", 1, data);
    # data = ["张三", 34]
    # eo.set_row("我的文档", None, data);
    # 搜索名字包含 "a" 的员工（不区分大小写）
    result = eo.search_in_column('Sheet1', 'name', '三')
    print(result)
